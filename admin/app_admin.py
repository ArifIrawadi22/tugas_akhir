"""
==============================================
  APLIKASI ADMIN - Port 5001
  Jalankan: python app_admin.py
  Buka: http://127.0.0.1:5001
==============================================
"""
from flask import Flask, render_template, request, redirect, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func
import pandas as pd
import numpy as np
import os, io
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'admin_rahasia_2024'
# Database SAMA — pakai path absolut agar admin & user baca DB yang sama
import os

DATABASE_URL = os.getenv("DATABASE_URL")

if DATABASE_URL:
    # Jika di Render → gunakan PostgreSQL
    app.config["SQLALCHEMY_DATABASE_URI"] = DATABASE_URL
else:
    # Jika di komputer lokal → gunakan SQLite
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    DB_PATH = os.path.join(BASE_DIR, "..", "kosmetik.db")
    app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_PATH}"

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# Konfigurasi upload gambar
UPLOAD_FOLDER = os.path.join(BASE_DIR, 'static', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
EKSTENSI_DIIZINKAN = {'png', 'jpg', 'jpeg', 'webp', 'jfif'}

def ekstensi_valid(nama_file):
    return '.' in nama_file and nama_file.rsplit('.', 1)[1].lower() in EKSTENSI_DIIZINKAN
# =====================
# MODEL DATABASE (sama dengan user)
# =====================
class Produk(db.Model):
    id          = db.Column(db.Integer, primary_key=True)
    brand       = db.Column(db.String(100))
    nama        = db.Column(db.String(200))
    tipe        = db.Column(db.String(50))
    ukuran      = db.Column(db.String(50))
    harga       = db.Column(db.Float)
    stok        = db.Column(db.Integer, default=0)
    terjual     = db.Column(db.Integer, default=0)
    rating      = db.Column(db.Float, default=0)
    bpom_id     = db.Column(db.String(50))
    url         = db.Column(db.String(300))
    deskripsi   = db.Column(db.Text)
    ingredients = db.Column(db.Text)
    gambar      = db.Column(db.String(200))

class Transaksi(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    produk_id  = db.Column(db.Integer, db.ForeignKey('produk.id'))
    jumlah     = db.Column(db.Integer)
    total      = db.Column(db.Float)
    tanggal    = db.Column(db.String(20))
    produk     = db.relationship('Produk', backref='transaksi')

class Pesanan(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    kode         = db.Column(db.String(30), unique=True)
    nama_pembeli = db.Column(db.String(100))
    email        = db.Column(db.String(100))
    no_hp        = db.Column(db.String(20))
    alamat       = db.Column(db.Text)
    total_harga  = db.Column(db.Float)
    status       = db.Column(db.String(20), default='Menunggu')
    status_bayar = db.Column(db.String(20), default='Belum Bayar') 
    tanggal      = db.Column(db.String(30))
    metode_bayar = db.Column(db.String(50), default='') 
    items        = db.relationship('PesananItem', backref='pesanan')

class PesananItem(db.Model):
    id         = db.Column(db.Integer, primary_key=True)
    pesanan_id = db.Column(db.Integer, db.ForeignKey('pesanan.id'))
    produk_id  = db.Column(db.Integer, db.ForeignKey('produk.id'))
    jumlah     = db.Column(db.Integer)
    harga_saat = db.Column(db.Float)
    produk     = db.relationship('Produk')


class Review(db.Model):
    id           = db.Column(db.Integer, primary_key=True)
    pesanan_id   = db.Column(db.Integer, db.ForeignKey('pesanan.id'))
    produk_id    = db.Column(db.Integer, db.ForeignKey('produk.id'))
    nama         = db.Column(db.String(100))
    rating       = db.Column(db.Integer)
    komentar     = db.Column(db.Text)
    tanggal      = db.Column(db.String(30))
    produk       = db.relationship('Produk', backref='reviews')
    pesanan      = db.relationship('Pesanan', backref='reviews')

# =====================
# IMPORT DATASET
# =====================
def import_dataset():
    if Produk.query.count() > 0:
        return
    path = os.path.join(BASE_DIR, '..', 'Product_Kosmetic.xlsx')
    if not os.path.exists(path):
        return
    df = pd.read_excel(path)
    np.random.seed(42)
    for _, row in df.iterrows():
        p = Produk(
            brand=str(row.get('brand','')), nama=str(row.get('product_name','')),
            tipe=str(row.get('product_type','')), ukuran=str(row.get('size','')),
            harga=float(row.get('normal_price',0)),
            stok=int(np.random.randint(10,100)), terjual=int(np.random.randint(5,80)),
            rating=round(float(np.random.uniform(3.5,5.0)),1),
            bpom_id=str(row.get('bpom_id','')), url=str(row.get('product_url','')),
            deskripsi=str(row.get('description_product','')),
            ingredients=str(row.get('ingredients_list',''))
        )
        db.session.add(p)
    db.session.commit()
    import random
    produk_list = Produk.query.all()
    # =====================
    # TRANSAKSI 2024
    # =====================

    bulan_2024 = [
        '2024-01','2024-02','2024-03','2024-04',
        '2024-05','2024-06','2024-07','2024-08',
        '2024-09','2024-10','2024-11','2024-12'
    ]

    for bulan in bulan_2024:
        for _ in range(random.randint(15,30)):

            p = random.choice(produk_list)
            qty = random.randint(1,5)

            db.session.add(
                Transaksi(
                    produk_id=p.id,
                    jumlah=qty,
                    total=qty * p.harga,
                    tanggal=bulan
                )
            )

    # =====================
    # TRANSAKSI 2025
    # =====================

    bulan_2025 = [
        '2025-01','2025-02','2025-03','2025-04',
        '2025-05','2025-06','2025-07','2025-08',
        '2025-09','2025-10','2025-11','2025-12'
    ]

    for bulan in bulan_2025:
        for _ in range(random.randint(20,40)):

            p = random.choice(produk_list)
            qty = random.randint(1,7)

            db.session.add(
                Transaksi(
                    produk_id=p.id,
                    jumlah=qty,
                    total=qty * p.harga,
                    tanggal=bulan
                )
            )

    # =====================
    # TRANSAKSI 2026
    # =====================

    bulan_2026 = [
        '2026-01',
        '2026-02',
        '2026-03',
        '2026-04',
        '2026-05'
    ]

    for bulan in bulan_2026:
        for _ in range(random.randint(10,25)):

            p = random.choice(produk_list)
            qty = random.randint(1,4)

            db.session.add(
                Transaksi(
                    produk_id=p.id,
                    jumlah=qty,
                    total=qty * p.harga,
                    tanggal=bulan
                )
            )

    db.session.commit()

def migrasi_kolom_gambar():
    """Tambah kolom 'gambar' ke tabel produk kalau database lama belum punya"""
    import sqlalchemy as sa
    inspector = sa.inspect(db.engine)
    if 'produk' in inspector.get_table_names():
        kolom_ada = [c['name'] for c in inspector.get_columns('produk')]
        if 'gambar' not in kolom_ada:
            with db.engine.connect() as conn:
                conn.execute(sa.text('ALTER TABLE produk ADD COLUMN gambar VARCHAR(200)'))
                conn.commit()
            print("✅ Migrasi: kolom 'gambar' berhasil ditambahkan ke database lama")
    if 'pesanan' in inspector.get_table_names():
        kolom_ada = [c['name'] for c in inspector.get_columns('pesanan')]
        if 'metode_bayar' not in kolom_ada:
            with db.engine.connect() as conn:
                conn.execute(sa.text("ALTER TABLE pesanan ADD COLUMN metode_bayar VARCHAR(50) DEFAULT ''"))
                conn.commit()
            print("✅ Migrasi: kolom 'metode_bayar' berhasil ditambahkan")
        if 'status_bayar' not in kolom_ada:
            with db.engine.connect() as conn:
                conn.execute(sa.text("ALTER TABLE pesanan ADD COLUMN status_bayar VARCHAR(20) DEFAULT 'Belum Bayar'"))
                conn.commit()
            print("✅ Migrasi: kolom 'status_bayar' berhasil ditambahkan")


def cek_login():
    return session.get('login', False)

# =====================
# ROUTES ADMIN
# =====================

@app.route('/', methods=['GET','POST'])
def login():
    error = ''
    if request.method == 'POST':
        if request.form.get('username')=='admin' and request.form.get('password')=='admin123':
            session['login'] = True
            return redirect('/dashboard')
        error = 'Username atau password salah!'
    return render_template('login.html', error=error)

@app.route('/logout')
def logout():
    session.clear(); return redirect('/')

@app.route('/dashboard')
def dashboard():
    if not cek_login():
        return redirect('/')

    tahun = request.args.get('tahun', '2024')

    total_produk  = Produk.query.count()

    total_brand   = db.session.query(
        func.count(func.distinct(Produk.brand))
    ).scalar()

    total_terjual = db.session.query(
        func.sum(Produk.terjual)
    ).scalar() or 0

    if tahun in ['2024', '2025']:
  
        total_omzet = db.session.query(
            func.sum(Transaksi.total)
        ).filter(
            Transaksi.tanggal.like(f'{tahun}-%')
        ).scalar() or 0

    else:

        omzet_dummy = db.session.query(
            func.sum(Transaksi.total)
        ).filter(
            Transaksi.tanggal.like(f'{tahun}-%')
        ).scalar() or 0

        # Hitung omzet real dari SEMUA pesanan tahun ini (semua status)
        # Format tanggal: "19 Jun 2026 03:33"
        omzet_real = 0
        for p in Pesanan.query.filter(
            Pesanan.tanggal.like(f'%{tahun}%')
        ).all():
            omzet_real += p.total_harga

        total_omzet = omzet_dummy + omzet_real
    stok_menipis  = Produk.query.filter(Produk.stok < 20).count()
    total_pesanan = Pesanan.query.count()
    menunggu      = Pesanan.query.filter_by(status='Menunggu').count()
    bulan_map = {'01':'Jan','02':'Feb','03':'Mar','04':'Apr','05':'Mei','06':'Jun',
                 '07':'Jul','08':'Ags','09':'Sep','10':'Okt','11':'Nov','12':'Des'}
    if tahun in ['2024', '2025']:
  
        rows = db.session.query(
            Transaksi.tanggal,
            func.sum(Transaksi.total)
        )\
        .filter(
            Transaksi.tanggal.like(f'{tahun}-%')
        )\
        .group_by(Transaksi.tanggal)\
        .order_by(Transaksi.tanggal)\
        .all()

    elif tahun == '2026':

        # Data dummy Jan-Mei 2026 dari tabel transaksi
        rows_dummy = db.session.query(
            Transaksi.tanggal,
            func.sum(Transaksi.total)
        )\
        .filter(
            Transaksi.tanggal.in_([
                '2026-01','2026-02','2026-03',
                '2026-04','2026-05'
            ])
        )\
        .group_by(Transaksi.tanggal)\
        .order_by(Transaksi.tanggal)\
        .all()

        # Data REAL Juni 2026 dari tabel pesanan
        # Format tanggal pesanan: "19 Jun 2026 03:33"
        # Ambil semua pesanan yang mengandung "Jun 2026" (semua status)
        semua_pesanan_juni = Pesanan.query.filter(
            Pesanan.tanggal.like('%Jun 2026%')
        ).all()

        total_juni_real = sum(p.total_harga for p in semua_pesanan_juni)

        rows = list(rows_dummy)

        # Selalu tampilkan Juni kalau ada pesanan, 
        # meski totalnya 0 tetap ditampilkan biar grafik konsisten
        if semua_pesanan_juni:
            rows.append(('2026-06', total_juni_real))
    bulan_label = [bulan_map.get(r[0].split('-')[1], r[0]) for r in rows]
    bulan_total = [int(r[1]) for r in rows]
    terlaris = Produk.query.order_by(Produk.terjual.desc()).limit(5).all()
    tipe_rows = db.session.query(Produk.tipe, func.count(Produk.id)).group_by(Produk.tipe).all()
    return render_template('dashboard.html',
        tahun=tahun,
        total_produk=total_produk, total_brand=total_brand,
        total_terjual=total_terjual, total_omzet=total_omzet,
        stok_menipis=stok_menipis, total_pesanan=total_pesanan,
        menunggu=menunggu, terlaris=terlaris,
        bulan_label=bulan_label, bulan_total=bulan_total,
        tipe_label=[r[0] for r in tipe_rows],
        tipe_count=[r[1] for r in tipe_rows])

@app.route('/produk')
def produk():
    if not cek_login(): return redirect('/')
    cari = request.args.get('cari','')
    tipe = request.args.get('tipe','')
    query = Produk.query
    if cari: query = query.filter(Produk.nama.ilike(f'%{cari}%')|Produk.brand.ilike(f'%{cari}%'))
    if tipe: query = query.filter(Produk.tipe==tipe)
    data = query.order_by(Produk.terjual.desc()).all()
    tipe_list = [r[0] for r in db.session.query(func.distinct(Produk.tipe)).all()]
    return render_template('produk.html', produk=data, tipe_list=tipe_list, cari=cari, tipe=tipe)

@app.route('/produk/tambah', methods=['POST'])
def tambah_produk():
    if not cek_login(): return redirect('/')

    nama_file_gambar = ''
    file = request.files.get('gambar')
    if file and file.filename != '' and ekstensi_valid(file.filename):
        from werkzeug.utils import secure_filename
        import uuid
        ext = file.filename.rsplit('.', 1)[1].lower()
        nama_file_gambar = f"{uuid.uuid4().hex}.{ext}"
        file.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_file_gambar))

    p = Produk(brand=request.form['brand'], nama=request.form['nama'],
               tipe=request.form['tipe'], ukuran=request.form['ukuran'],
               harga=float(request.form['harga']), stok=int(request.form['stok']),
               terjual=0, rating=0, bpom_id=request.form.get('bpom_id',''),
               deskripsi=request.form.get('deskripsi',''), ingredients='', url='',
               gambar=nama_file_gambar)
    db.session.add(p); db.session.commit()
    return redirect('/produk')

@app.route('/produk/hapus/<int:id>')
def hapus_produk(id):
    if not cek_login(): return redirect('/')
    p = Produk.query.get_or_404(id)
    # Hapus juga file gambar dari folder kalau ada
    if p.gambar:
        path_gambar = os.path.join(app.config['UPLOAD_FOLDER'], p.gambar)
        if os.path.exists(path_gambar):
            os.remove(path_gambar)
    db.session.delete(p); db.session.commit()
    return redirect('/produk')

@app.route('/produk/edit/<int:id>', methods=['GET','POST'])
def edit_produk(id):
    if not cek_login(): return redirect('/')
    p = Produk.query.get_or_404(id)
    if request.method == 'POST':
        p.brand=request.form['brand']; p.nama=request.form['nama']
        p.tipe=request.form['tipe']; p.ukuran=request.form['ukuran']
        p.harga=float(request.form['harga']); p.stok=int(request.form['stok'])
        p.bpom_id=request.form.get('bpom_id','')
        p.deskripsi=request.form.get('deskripsi','')

        # Ganti gambar kalau user upload file baru
        file = request.files.get('gambar')
        if file and file.filename != '' and ekstensi_valid(file.filename):
            import uuid
            # Hapus gambar lama kalau ada
            if p.gambar:
                path_lama = os.path.join(app.config['UPLOAD_FOLDER'], p.gambar)
                if os.path.exists(path_lama):
                    os.remove(path_lama)
            ext = file.filename.rsplit('.', 1)[1].lower()
            nama_file_baru = f"{uuid.uuid4().hex}.{ext}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'], nama_file_baru))
            p.gambar = nama_file_baru

        # Hapus gambar kalau user centang "hapus gambar"
        if request.form.get('hapus_gambar') == 'on' and p.gambar:
            path_lama = os.path.join(app.config['UPLOAD_FOLDER'], p.gambar)
            if os.path.exists(path_lama):
                os.remove(path_lama)
            p.gambar = None

        db.session.commit(); return redirect('/produk')
    return render_template('edit_produk.html', produk=p)

@app.route('/laporan')
def laporan():
    if not cek_login(): return redirect('/')
    tipe_rows = db.session.query(Produk.tipe, func.sum(Transaksi.total), func.sum(Transaksi.jumlah))\
                    .join(Transaksi, Produk.id==Transaksi.produk_id).group_by(Produk.tipe).all()
    brand_rows = db.session.query(Produk.brand, func.sum(Transaksi.total))\
                    .join(Transaksi, Produk.id==Transaksi.produk_id)\
                    .group_by(Produk.brand).order_by(func.sum(Transaksi.total).desc()).limit(10).all()
    transaksi = Transaksi.query.order_by(Transaksi.id.desc()).limit(20).all()
    return render_template('laporan.html', tipe_rows=tipe_rows,
                           brand_rows=brand_rows, transaksi=transaksi)

@app.route('/laporan/export')
def export_excel():
    if not cek_login(): return redirect('/')
    produk = Produk.query.all()
    data = [{'Brand':p.brand,'Nama Produk':p.nama,'Tipe':p.tipe,
             'Ukuran':p.ukuran,'Harga':p.harga,'Stok':p.stok,
             'Terjual':p.terjual,'Rating':p.rating,'BPOM ID':p.bpom_id,
             'Total Omzet':p.harga*p.terjual} for p in produk]
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as w:
        pd.DataFrame(data).to_excel(w, index=False, sheet_name='Produk')
    output.seek(0)
    return send_file(output, download_name='laporan_kosmetik.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/pesanan')
def pesanan():
    if not cek_login(): return redirect('/')
    status_filter = request.args.get('status','')
    query = Pesanan.query
    if status_filter: query = query.filter(Pesanan.status==status_filter)
    pesanan_list = query.order_by(Pesanan.id.desc()).all()
    return render_template('admin_pesanan.html', pesanan_list=pesanan_list,
        status_filter=status_filter,
        total_pesanan=Pesanan.query.count(),
        total_menunggu=Pesanan.query.filter_by(status='Menunggu').count(),
        total_diproses=Pesanan.query.filter_by(status='Diproses').count(),
        total_selesai=Pesanan.query.filter_by(status='Selesai').count())

@app.route('/pesanan/export')
def export_pesanan():
    if not cek_login(): return redirect('/')
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime as dt

    pesanan_list = Pesanan.query.order_by(Pesanan.id.desc()).all()

    wb = openpyxl.Workbook()

    # ── Warna ──
    UNGU  = "7C3AED"; PINK  = "E879A0"; HIJAU = "15803D"
    PUTIH = "FFFFFF"; ABU   = "F3F4F6"; UNGU_L= "EDE9FE"
    MERAH = "991B1B"; BIRU  = "1D4ED8"; KUNING= "D97706"

    def hdr(cell, teks, bg=UNGU, fg=PUTIH, bold=True, size=10):
        cell.value = teks
        cell.font = Font(name='Arial', bold=bold, color=fg, size=size)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def border_all(ws, r1, r2, c1, c2):
        thin = Side(style='thin', color='D1D5DB')
        for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2):
            for cell in row:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ══════════════════════════════════════
    # SHEET 1 — DATA PESANAN
    # ══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Data Pesanan"

    ws1.merge_cells('A1:K1')
    hdr(ws1['A1'], f'📦 DATA PESANAN MASUK — Toko Kosmetik  (Export: {dt.now().strftime("%d %b %Y %H:%M")})',
        UNGU, PUTIH, True, 13)
    ws1.row_dimensions[1].height = 32

    headers = ['No','Kode Pesanan','Nama Pembeli','Email','No HP',
               'Produk','Jumlah','Total (Rp)','Status','Metode Pembayaran','Status Bayar','Tanggal']
    for i, h in enumerate(headers, 1):
        hdr(ws1.cell(2, i), h, UNGU, PUTIH)
    ws1.row_dimensions[2].height = 25

    warna_status = {
        'Menunggu': 'FEF9C3', 'Diproses': 'DBEAFE',
        'Dikirim':  'E0F2FE', 'Selesai':  'DCFCE7', 'Dibatalkan': 'FEE2E2'
    }
    warna_teks = {
        'Menunggu': KUNING, 'Diproses': BIRU,
        'Dikirim':  '0369A1', 'Selesai': HIJAU, 'Dibatalkan': MERAH
    }

    for idx, p in enumerate(pesanan_list, 1):
        r = idx + 2
        ws1.row_dimensions[r].height = 18
        nama_produk = ', '.join([
            f"{item.produk.nama[:30] if item.produk else '-'} ({item.jumlah}x)"
            for item in p.items
        ])
        bg = ABU if idx % 2 == 0 else PUTIH
        data = [idx, p.kode, p.nama_pembeli, p.email, p.no_hp,
                nama_produk, sum(i.jumlah for i in p.items),
                p.total_harga,p.status, p.metode_bayar, p.status_bayar, p.tanggal]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(r, col)
            cell.value = val
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(
                horizontal='center' if col in [1,7,9] else 'left',
                vertical='center', wrap_text=(col==6)
            )
            cell.fill = PatternFill('solid', start_color=bg)
            if col == 8:
                cell.number_format = '#,##0'
                cell.font = Font(name='Arial', size=9, color=HIJAU, bold=True)
            if col == 9:
                sbg = warna_status.get(str(val), ABU)
                stxt = warna_teks.get(str(val), '374151')
                cell.fill = PatternFill('solid', start_color=sbg)
                cell.font = Font(name='Arial', size=9, color=stxt, bold=True)

    # Baris total
    total_r = len(pesanan_list) + 3
    ws1.merge_cells(f'A{total_r}:G{total_r}')
    ws1.cell(total_r,1).value = f'TOTAL — {len(pesanan_list)} Pesanan'
    ws1.cell(total_r,1).font = Font(name='Arial', bold=True, color=UNGU)
    ws1.cell(total_r,1).fill = PatternFill('solid', start_color=UNGU_L)
    ws1.cell(total_r,1).alignment = Alignment(horizontal='center')
    ws1.cell(total_r,8).value = f'=SUM(H3:H{total_r-1})'
    ws1.cell(total_r,8).number_format = '#,##0'
    ws1.cell(total_r,8).font = Font(name='Arial', bold=True, color=HIJAU, size=11)
    ws1.cell(total_r,8).fill = PatternFill('solid', start_color=UNGU_L)
    for col in [9,10]:
        ws1.cell(total_r,col).fill = PatternFill('solid', start_color=UNGU_L)

    border_all(ws1, 1, total_r, 1, 10)
    for col, w in zip(range(1,11),[5,22,18,26,15,40,9,16,12,16]):
        ws1.column_dimensions[get_column_letter(col)].width = w
    ws1.freeze_panes = 'A3'

    # ══════════════════════════════════════
    # SHEET 2 — ULASAN & RATING
    # ══════════════════════════════════════
    ws2 = wb.create_sheet("Ulasan & Rating")

    ws2.merge_cells('A1:G1')
    hdr(ws2['A1'], f'⭐ DATA ULASAN & RATING PRODUK  (Export: {dt.now().strftime("%d %b %Y %H:%M")})',
        PINK, PUTIH, True, 13)
    ws2.row_dimensions[1].height = 32

    rev_headers = ['No','Nama Pembeli','Produk','Kategori','Rating (⭐)','Komentar','Tanggal']
    for i, h in enumerate(rev_headers, 1):
        hdr(ws2.cell(2, i), h, PINK, PUTIH)
    ws2.row_dimensions[2].height = 25

    semua_review = Review.query.order_by(Review.id.desc()).all()
    for idx, rv in enumerate(semua_review, 1):
        r = idx + 2
        ws2.row_dimensions[r].height = 20
        bg = ABU if idx % 2 == 0 else PUTIH
        bintang = '★' * rv.rating + '☆' * (5 - rv.rating)
        data = [idx, rv.nama,
                rv.produk.nama[:50] if rv.produk else '-',
                rv.produk.tipe if rv.produk else '-',
                f"{rv.rating}/5  {bintang}",
                rv.komentar, rv.tanggal]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(r, col)
            cell.value = val
            cell.font = Font(name='Arial', size=9)
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='center' if col in [1,4,5] else 'left',
                                       vertical='center', wrap_text=(col==6))
            if col == 5:
                warna_rating = {5:HIJAU, 4:'15803D', 3:KUNING, 2:KUNING, 1:MERAH}
                cell.font = Font(name='Arial', size=9, bold=True,
                                color=warna_rating.get(rv.rating, '374151'))

    # Ringkasan rating di bawah
    if semua_review:
        gap = len(semua_review) + 5
        ws2.merge_cells(f'A{gap}:G{gap}')
        hdr(ws2.cell(gap,1), '📊 Ringkasan Rating', PINK, PUTIH, True, 11)
        ws2.row_dimensions[gap].height = 28

        for i, bintang in enumerate([5,4,3,2,1], 1):
            r = gap + i
            jml = sum(1 for rv in semua_review if rv.rating == bintang)
            pct = jml/len(semua_review)*100 if semua_review else 0
            ws2.cell(r,1).value = f'{"★"*bintang}'
            ws2.cell(r,1).font = Font(name='Arial', size=11, color='F59E0B')
            ws2.cell(r,2).value = jml
            ws2.cell(r,2).font = Font(name='Arial', bold=True)
            ws2.cell(r,3).value = f'{pct:.1f}%'
            ws2.cell(r,3).font = Font(name='Arial', size=9, color='6B7280')
            for col in range(1,4):
                ws2.cell(r,col).fill = PatternFill('solid', start_color=ABU if i%2==0 else PUTIH)
                ws2.cell(r,col).alignment = Alignment(horizontal='center')

        avg_r = gap + 7
        avg_rating = sum(rv.rating for rv in semua_review) / len(semua_review)
        ws2.merge_cells(f'A{avg_r}:C{avg_r}')
        ws2.cell(avg_r,1).value = f'Rata-rata Rating: {avg_rating:.1f} / 5.0  ⭐'
        ws2.cell(avg_r,1).font = Font(name='Arial', bold=True, size=12, color=UNGU)
        ws2.cell(avg_r,1).fill = PatternFill('solid', start_color=UNGU_L)
        ws2.cell(avg_r,1).alignment = Alignment(horizontal='center')

    border_all(ws2, 1, len(semua_review)+2, 1, 7)
    for col, w in zip(range(1,8),[5,20,40,14,14,45,14]):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.freeze_panes = 'A3'

    # ══════════════════════════════════════
    # SHEET 3 — RINGKASAN STATUS
    # ══════════════════════════════════════
    ws3 = wb.create_sheet("Ringkasan Status")

    ws3.merge_cells('A1:D1')
    hdr(ws3['A1'], '📋 RINGKASAN STATUS PESANAN', UNGU, PUTIH, True, 13)
    ws3.row_dimensions[1].height = 32

    for i, h in enumerate(['Status','Jumlah Pesanan','Total Omzet (Rp)','Persentase'], 1):
        hdr(ws3.cell(2,i), h, UNGU, PUTIH)
    ws3.row_dimensions[2].height = 25

    status_list = ['Menunggu','Diproses','Dikirim','Selesai','Dibatalkan']
    for idx, st in enumerate(status_list, 1):
        r = idx + 2
        items = [p for p in pesanan_list if p.status == st]
        total = sum(p.total_harga for p in items)
        sbg = warna_status.get(st, ABU)
        stxt = warna_teks.get(st, '374151')
        ws3.cell(r,1).value = st
        ws3.cell(r,1).font = Font(name='Arial', bold=True, color=stxt, size=10)
        ws3.cell(r,1).fill = PatternFill('solid', start_color=sbg)
        ws3.cell(r,1).alignment = Alignment(horizontal='center')
        ws3.cell(r,2).value = len(items)
        ws3.cell(r,2).font = Font(name='Arial', bold=True, size=11, color=stxt)
        ws3.cell(r,2).fill = PatternFill('solid', start_color=sbg)
        ws3.cell(r,2).alignment = Alignment(horizontal='center')
        ws3.cell(r,3).value = total
        ws3.cell(r,3).number_format = '#,##0'
        ws3.cell(r,3).font = Font(name='Arial', size=10, color=HIJAU, bold=True)
        ws3.cell(r,3).fill = PatternFill('solid', start_color=sbg)
        ws3.cell(r,3).alignment = Alignment(horizontal='center')
        ws3.cell(r,4).value = f'=B{r}/SUM(B3:B7)' if len(pesanan_list)>0 else 0
        ws3.cell(r,4).number_format = '0.0%'
        ws3.cell(r,4).fill = PatternFill('solid', start_color=sbg)
        ws3.cell(r,4).alignment = Alignment(horizontal='center')
        ws3.row_dimensions[r].height = 22

    total_row = 8
    for i, h in enumerate(['TOTAL', f'=SUM(B3:B7)', f'=SUM(C3:C7)', '100%'], 1):
        ws3.cell(total_row,i).value = h
        ws3.cell(total_row,i).fill = PatternFill('solid', start_color=UNGU_L)
        ws3.cell(total_row,i).font = Font(name='Arial', bold=True, color=UNGU, size=11)
        ws3.cell(total_row,i).alignment = Alignment(horizontal='center')
        if i==3: ws3.cell(total_row,i).number_format = '#,##0'

    border_all(ws3, 1, total_row, 1, 4)
    for col, w in zip(range(1,5),[16,20,22,14]):
        ws3.column_dimensions[get_column_letter(col)].width = w

    # Kirim file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    tgl = dt.now().strftime('%Y%m%d_%H%M')
    return send_file(output,
                     download_name=f'laporan_pesanan_{tgl}.xlsx',
                     as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/pesanan/update/<int:id>/<status>')
def update_pesanan(id, status):
    if not cek_login(): return redirect('/')
    p = Pesanan.query.get_or_404(id)
    if status in ['Menunggu','Diproses','Dikirim','Selesai','Dibatalkan']:
        p.status = status; db.session.commit()
    return redirect('/pesanan')

@app.route('/pesanan/konfirmasi-bayar/<int:id>')
def konfirmasi_bayar(id):
    """Admin konfirmasi pembayaran sudah diterima"""
    if not cek_login(): return redirect('/')
    p = Pesanan.query.get_or_404(id)
    p.status_bayar = 'Sudah Bayar'
    # Otomatis ubah status pesanan jadi Diproses kalau masih Menunggu
    if p.status == 'Menunggu':
        p.status = 'Diproses'
    db.session.commit()
    return redirect('/pesanan')

@app.route('/database')
def lihat_database():
    if not cek_login(): return redirect('/')
    tabel = request.args.get('tabel','produk')
    data, kolom, total = [], [], 0
    if tabel=='produk':
        items = Produk.query.all()
        kolom = ['ID','Brand','Nama','Tipe','Harga','Stok','Terjual','Rating','BPOM']
        data  = [[p.id,p.brand,p.nama[:40],p.tipe,f"Rp {p.harga:,.0f}",
                  p.stok,p.terjual,p.rating,p.bpom_id] for p in items]
        total = len(items)
    elif tabel=='pesanan':
        items = Pesanan.query.order_by(Pesanan.id.desc()).all()
        kolom = ['ID','Kode','Pembeli','Email','Total','Status','Metode Pembayaran','Tanggal']
        data  = [[p.id,p.kode,p.nama_pembeli,p.email,
                  f"Rp {p.total_harga:,.0f}",p.status,p.tanggal] for p in items]
        total = len(items)
    elif tabel=='transaksi':
        items = Transaksi.query.order_by(Transaksi.id.desc()).all()
        kolom = ['ID','Produk','Jumlah','Total','Bulan']
        data  = [[t.id, t.produk.nama[:35] if t.produk else '-',
                  t.jumlah, f"Rp {t.total:,.0f}", t.tanggal] for t in items]
        total = len(items)
    return render_template('lihat_database.html', tabel=tabel,
                           kolom=kolom, data=data, total=total)
    
@app.route('/database/export')
def export_database():

    if not cek_login():
        return redirect('/')

    produk = Produk.query.all()

    data = []

    for p in produk:
        data.append({
            'ID': p.id,
            'Brand': p.brand,
            'Nama Produk': p.nama,
            'Tipe': p.tipe,
            'Ukuran': p.ukuran,
            'Harga': p.harga,
            'Stok': p.stok,
            'Terjual': p.terjual,
            'Rating': p.rating,
            'BPOM ID': p.bpom_id
        })

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        pd.DataFrame(data).to_excel(
            writer,
            sheet_name='Database Produk',
            index=False
        )

    output.seek(0)

    return send_file(
        output,
        download_name='database_produk.xlsx',
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# =====================
# ROUTE MACHINE LEARNING
# =====================
@app.route('/ml')
def ml_dashboard():
    if not cek_login(): return redirect('/')
    import sys, os
    sys.path.insert(0, os.path.dirname(__file__))
    import importlib
    import ml_engine as mle
    importlib.reload(mle)

    tab = request.args.get('tab', 'rekomendasi')

    # Data untuk semua produk (dropdown)
    semua_produk = Produk.query.order_by(Produk.terjual.desc()).all()
    produk_id = request.args.get('produk_id', type=int)

    rekomendasi = []
    if produk_id:
        rekomendasi = mle.rekomendasi_produk(produk_id, top=6)

    return render_template('ml_dashboard.html',
        tab=tab,
        semua_produk=semua_produk,
        produk_id=produk_id,
        rekomendasi=rekomendasi,
        sentimen_data=mle.ringkasan_sentimen_produk() if tab=='sentimen' else {'total':0,'positif':0,'netral':0,'negatif':0,'detail':[]},
        clustering_data=mle.clustering_produk() if tab=='clustering' else {'produk':[],'stats':[]},
        prediksi_data=mle.prediksi_penjualan(3) if tab=='prediksi' else {'historis':[],'prediksi_depan':[],'akurasi':0,'r2':0,'mae':0},
        rdf_data=mle.buat_knowledge_graph() if tab=='rdf' else {'triple_count':0,'produk_count':0,'turtle_preview':'','turtle_full':'','graph_nodes':[],'graph_edges':[]},
        stok_data=mle.prediksi_stok_produk() if tab=='stok' else [],
        rincian_bulan=mle.rincian_penjualan_per_bulan() if tab=='rincian' else []
    )

@app.route('/ml/rdf/download')
def download_rdf():
    if not cek_login(): return redirect('/')
    import sys, os
    sys.path.insert(0, os.path.dirname(__file__))
    import ml_engine as mle
    kg = mle.buat_knowledge_graph()
    output = io.BytesIO(kg['turtle_full'].encode('utf-8'))
    output.seek(0)
    return send_file(output, download_name='produk_kosmetik.ttl',
                     as_attachment=True, mimetype='text/turtle')

from flask import Response
@app.route('/rdf')
def lihat_rdf():

    if not cek_login():
        return redirect('/')

    import sys, os
    sys.path.insert(0, os.path.dirname(__file__))

    import ml_engine as mle

    kg = mle.buat_knowledge_graph()

    return Response(
        kg['turtle_full'],
        mimetype='text/turtle'
    )
    
with app.app_context():
    db.create_all()

if __name__ == '__main__':
    migrasi_kolom_gambar()
    import_dataset()
    print("=" * 45)
    print("  ADMIN PANEL berjalan di port 5001")
    print("  Buka: http://127.0.0.1:5001")
    print("  Login: admin / admin123")
    print("=" * 45)
    app.run(debug=True, port=5001)
